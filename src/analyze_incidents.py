import argparse
from pathlib import Path
import pandas as pd

from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference


SLA_THRESHOLDS = {"P1": 60, "P2": 240, "P3": 1440, "P4": 2880}


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 50)


def style_header(ws, row=1):
    header_fill = PatternFill("solid", fgColor="1F2937")  # dark
    header_font = Font(color="FFFFFF", bold=True)
    for c in ws[row]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")


def write_df(ws, df, start_row=1, start_col=1):
    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, v in enumerate(r, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=v)
    style_header(ws, start_row)
    autosize(ws)


def main():
    parser = argparse.ArgumentParser(description="Endpoint + Security Incident Trend Analyzer: CSV -> Excel report")
    parser.add_argument("csv_path", help="Path to incident CSV export")
    parser.add_argument("--out", default="reports/incident_trends_report.xlsx", help="Output Excel report path")
    args = parser.parse_args()

    csv_path = Path(args.csv_path)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    df = pd.read_csv(csv_path)

    required = [
        "incident_id","opened_at","resolved_at","user_role","device_type","site",
        "network_path","vendor","issue_category","priority","resolution_minutes","resolved"
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    # Normalize
    df["opened_at"] = pd.to_datetime(df["opened_at"], errors="coerce")
    df["resolved_at"] = pd.to_datetime(df["resolved_at"], errors="coerce")
    df["resolved"] = df["resolved"].astype(str).str.strip().str.lower().map({"yes": True, "no": False})
    df["priority"] = df["priority"].astype(str).str.strip()
    df["resolution_minutes"] = pd.to_numeric(df["resolution_minutes"], errors="coerce")

    df["date"] = df["opened_at"].dt.date
    df["is_executive"] = df["user_role"].astype(str).str.strip().str.lower().eq("executive")

    df["sla_minutes"] = df["priority"].map(SLA_THRESHOLDS)
    df["sla_breached"] = (
        df["resolution_minutes"].notna()
        & df["sla_minutes"].notna()
        & (df["resolution_minutes"] > df["sla_minutes"])
    )

    total = len(df)
    resolved_count = int(df["resolved"].fillna(False).sum())
    unresolved_count = total - resolved_count

    mttr = df.loc[df["resolution_minutes"].notna(), "resolution_minutes"].mean()
    p95 = df.loc[df["resolution_minutes"].notna(), "resolution_minutes"].quantile(0.95)

    # Trends (daily)
    trends = (
        df.groupby("date", dropna=False)
        .agg(
            incidents=("incident_id","count"),
            resolved=("resolved", lambda s: int(s.fillna(False).sum())),
            unresolved=("resolved", lambda s: int((~s.fillna(False)).sum()))
        )
        .reset_index()
        .sort_values("date")
    )

    # RCA-style category breakdown
    by_cat = (
        df.groupby("issue_category")
        .agg(
            count=("incident_id","count"),
            pct=("incident_id", lambda s: round(len(s) / max(total,1) * 100, 1)),
            avg_minutes=("resolution_minutes","mean"),
            median_minutes=("resolution_minutes","median"),
            max_minutes=("resolution_minutes","max"),
            breaches=("sla_breached", lambda s: int(s.sum()))
        )
        .reset_index()
        .sort_values(["count","breaches"], ascending=[False, False])
    )
    for c in ["avg_minutes","median_minutes","max_minutes"]:
        by_cat[c] = by_cat[c].round(1)

    # Executive impact
    exec_impact = (
        df.groupby(["is_executive","issue_category"])
        .size()
        .reset_index(name="count")
        .assign(user_group=lambda x: x["is_executive"].map({True:"Executive", False:"Non-Executive"}))
        .drop(columns=["is_executive"])
        .sort_values(["user_group","count"], ascending=[True, False])
    )

    # SLA risk
    sla = (
        df[df["sla_minutes"].notna() & df["resolution_minutes"].notna()]
        .groupby("priority")
        .agg(total_with_sla=("incident_id","count"), breaches=("sla_breached", lambda s: int(s.sum())))
        .reset_index()
    )
    sla["breach_rate_percent"] = (sla["breaches"] / sla["total_with_sla"] * 100).round(1)

    # Heatmap-like pivot (Category x Priority)
    heat = pd.pivot_table(
        df, index="issue_category", columns="priority", values="incident_id",
        aggfunc="count", fill_value=0
    ).reset_index()

    # Recommendations (simple, credible)
    top = by_cat.head(5)["issue_category"].tolist()
    recs = []
    for issue in top:
        i = issue.lower()
        if "vpn" in i or "remote access" in i:
            recs.append("Standardize remote connectivity: client versions, certificates, MFA/token health, and VPN profiles.")
        elif "o365" in i:
            recs.append("Review O365 sign-in failures: conditional access, MFA policies, and identity provider health.")
        elif "teams" in i:
            recs.append("Reduce Teams incidents: baseline AV drivers/firmware, device profile standards, and known-good configs.")
        elif "conference" in i or "av" in i:
            recs.append("Conference room reliability: pre-meeting health checks + standard room profiles + vendor runbooks.")
        elif "edr" in i or "phishing" in i or "mfa" in i:
            recs.append("Security workflow: triage playbooks + alert classification + escalation paths + SLA-backed response.")
        else:
            recs.append(f"{issue}: create a repeatable fix playbook and measure post-change incident reduction.")
    recommendations = pd.DataFrame({"recommendation": recs})

    # Executive summary table
    summary = pd.DataFrame(
        [
            ("Total incidents", total),
            ("Resolved incidents", resolved_count),
            ("Unresolved incidents", unresolved_count),
            ("Resolution rate (%)", round(resolved_count / max(total,1) * 100, 1)),
            ("MTTR (minutes)", round(mttr, 1) if pd.notna(mttr) else "N/A"),
            ("P95 resolution (minutes)", round(p95, 1) if pd.notna(p95) else "N/A"),
            ("Top issue category", by_cat.iloc[0]["issue_category"] if len(by_cat) else "N/A"),
        ],
        columns=["metric","value"]
    )

    # Write Excel with formatting + charts
    from openpyxl import Workbook
    wb = Workbook()

    # Replace default sheet
    wb.remove(wb.active)

    ws = wb.create_sheet("Executive_Summary")
    write_df(ws, summary)

    ws2 = wb.create_sheet("Trends_Daily")
    write_df(ws2, trends)

    # Line chart (incidents over time)
    if len(trends) >= 2:
        chart = LineChart()
        chart.title = "Daily Incident Volume"
        chart.y_axis.title = "Incidents"
        chart.x_axis.title = "Date"

        data = Reference(ws2, min_col=2, min_row=1, max_col=2, max_row=1 + len(trends))
        cats = Reference(ws2, min_col=1, min_row=2, max_row=1 + len(trends))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 22
        ws2.add_chart(chart, "E2")

    ws3 = wb.create_sheet("Category_RCA")
    write_df(ws3, by_cat)

    # Bar chart top 10 categories
    top_n = min(10, len(by_cat))
    if top_n >= 3:
        chart2 = BarChart()
        chart2.title = "Top Issue Categories (Count)"
        chart2.y_axis.title = "Count"
        chart2.x_axis.title = "Category"

        # Write a small chart table to the right
        chart_df = by_cat[["issue_category","count"]].head(top_n).copy()
        start_row = 2
        start_col = 10
        ws3.cell(1, start_col, "issue_category")
        ws3.cell(1, start_col+1, "count")
        ws3.cell(1, start_col).font = Font(bold=True)
        ws3.cell(1, start_col+1).font = Font(bold=True)

        for i, row in enumerate(chart_df.itertuples(index=False), start=start_row):
            ws3.cell(i, start_col, row.issue_category)
            ws3.cell(i, start_col+1, int(row.count))

        data = Reference(ws3, min_col=start_col+1, min_row=1, max_row=start_row + top_n - 1)
        cats = Reference(ws3, min_col=start_col, min_row=2, max_row=start_row + top_n - 1)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.height = 12
        chart2.width = 24
        ws3.add_chart(chart2, "J2")

    ws4 = wb.create_sheet("SLA_Risk")
    write_df(ws4, sla)

    ws5 = wb.create_sheet("Executive_Impact")
    write_df(ws5, exec_impact)

    ws6 = wb.create_sheet("Category_x_Priority")
    write_df(ws6, heat)

    ws7 = wb.create_sheet("Recommendations")
    write_df(ws7, recommendations)

    ws8 = wb.create_sheet("Raw_Data")
    write_df(ws8, df.sort_values("opened_at"))

    wb.save(out_path)

    print("Endpoint + Security Incident Trend Analyzer")
    print(f"Input:  {csv_path}")
    print(f"Output: {out_path}")
    print("\nTop 5 categories:")
    print(by_cat[["issue_category","count","avg_minutes","breaches"]].head(5).to_string(index=False))
    print("\nKey KPIs:")
    for m, v in summary.values.tolist():
        print(f"- {m}: {v}")


if __name__ == "__main__":
    main()
